import streamlit as st
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage # NameError 방지
import easyocr
import io

# 1. 페이지 설정
st.set_page_config(page_title="번호판 인식기", layout="centered")
st.title("🚗 번호판 인식 및 엑셀 자동 배치")
st.write("이미지를 업로드하면 가로 11cm, 세로 13cm 크기로 엑셀에 저장됩니다.")

# OCR 모델 로드 (캐싱하여 속도 향상)
@st.cache_resource
def load_ocr():
    # 로컬에서 처음 실행 시 모델 다운로드로 인해 시간이 걸릴 수 있습니다.
    return easyocr.Reader(['ko', 'en'])

try:
    reader = load_ocr()
except Exception as e:
    st.error(f"OCR 모델 로드 중 오류: {e}")

# 2. 파일 업로드
uploaded_file = st.file_uploader("번호판 사진 업로드", type=['jpg', 'jpeg', 'png'])

if uploaded_file is not None:
    # 사진 미리보기
    st.image(uploaded_file, caption="업로드된 이미지", use_container_width=True)
    
    with st.spinner("번호 분석 중..."):
        try:
            image_bytes = uploaded_file.read()
            results = reader.readtext(image_bytes)
            result_text = " ".join([res[1] for res in results])
            st.success(f"인식 결과: {result_text}")
        except Exception as e:
            st.error(f"이미지 분석 중 오류: {e}")
            result_text = "인식 실패"

    # 3. 엑셀 작업
    try:
        # 템플릿 로드 (test.xlsx가 같은 경로에 있어야 함)
        wb = load_workbook("test.xlsx")
        ws = wb.active

        # 이미지 객체 생성 및 크기 설정 (cm -> px 변환)
        img_for_excel = OpenpyxlImage(io.BytesIO(image_bytes))
        
        # 1cm = 약 37.8 픽셀 (엑셀 표준)
        cm_to_px = 37.8
        img_for_excel.width = 11 * cm_to_px  # 가로 11cm
        img_for_excel.height = 13 * cm_to_px # 세로 13cm
        
        # A3 셀에 배치 (VBA 없이 파이썬이 직접 수행)
        img_for_excel.anchor = 'A3'
        ws.add_image(img_for_excel)

        # A38 셀에 결과 입력
        ws['A38'] = result_text

        # 4. 파일 다운로드 준비
        excel_out = io.BytesIO()
        wb.save(excel_out)
        excel_out.seek(0)

        st.download_button(
            label="📊 결과 엑셀 다운로드",
            data=excel_out,
            file_name=f"result_{result_text}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except FileNotFoundError:
        st.error("오류: 'test.xlsx' 파일을 찾을 수 없습니다. 파일이 같은 폴더에 있는지 확인하세요.")
    except Exception as e:
        st.error(f"엑셀 생성 중 오류 발생: {e}")